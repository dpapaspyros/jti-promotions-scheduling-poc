"""Tests for the XLSX schedule export endpoint."""

import io
from datetime import date

import openpyxl
from rest_framework.test import APITestCase

from ._helpers import _make_pos, _make_promoter, _make_schedule, _make_user, _make_visit


class ScheduleExportTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.pos = _make_pos("POS001", "Test POS", city="Athens")
        self.promoter = _make_promoter("alice", "Alice", "Smith")
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30)
        )

    def _get(self, pk=None):
        return self.client.get(f"/api/schedules/{pk or self.schedule.pk}/export/")

    def _load_wb(self, response):
        return openpyxl.load_workbook(io.BytesIO(response.content))

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(self._get().status_code, 401)

    def test_not_found_returns_404(self):
        self.assertEqual(self._get(pk=99999).status_code, 404)

    def test_returns_xlsx_content_type(self):
        response = self._get()
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])

    def test_content_disposition_includes_schedule_name(self):
        response = self._get()
        self.assertIn("April 2026", response["Content-Disposition"])
        self.assertIn("attachment", response["Content-Disposition"])

    def test_xlsx_header_row(self):
        headers = [cell.value for cell in self._load_wb(self._get()).active[1]]
        self.assertIn("CDB Code", headers)
        self.assertIn("POS Name", headers)
        self.assertIn("Promoter", headers)

    def test_xlsx_contains_visit_data(self):
        _make_visit(self.schedule, self.pos, self.promoter, comments="Peak window.")
        rows = list(
            self._load_wb(self._get()).active.iter_rows(min_row=2, values_only=True)
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][4], "POS001")  # CDB Code column
        self.assertIn("Alice", rows[0][8])  # Promoter column

    def test_empty_schedule_returns_header_row_only(self):
        rows = list(
            self._load_wb(self._get()).active.iter_rows(min_row=2, values_only=True)
        )
        self.assertEqual(len(rows), 0)

    def test_visits_ordered_by_date_then_time(self):
        _make_visit(
            self.schedule,
            self.pos,
            visit_date=date(2026, 4, 10),
            start_time="15:00",
            end_time="17:00",
        )
        _make_visit(
            self.schedule,
            self.pos,
            visit_date=date(2026, 4, 3),
            start_time="09:00",
            end_time="11:00",
        )
        rows = list(
            self._load_wb(self._get()).active.iter_rows(min_row=2, values_only=True)
        )
        self.assertEqual(rows[0][1], "2026-04-03")
        self.assertEqual(rows[1][1], "2026-04-10")
