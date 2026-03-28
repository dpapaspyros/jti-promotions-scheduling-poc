"""XLSX export and import views."""

import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import PointOfSale, Promoter, Schedule, ScheduledVisit
from ..serializers import ScheduledVisitSerializer
from ._helpers import (
    _XLSX_COL_WIDTHS,
    _XLSX_HEADERS,
    _parse_xlsx_date,
    _parse_xlsx_time,
)


class ScheduleExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        schedule = get_object_or_404(Schedule, pk=pk)
        visits = (
            ScheduledVisit.objects.filter(schedule=schedule)
            .select_related("pos", "promoter")
            .order_by("date", "start_time")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = schedule.name[:31]  # Excel sheet name limit
        ws.freeze_panes = "A2"

        ws.append(_XLSX_HEADERS)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        for col, width in enumerate(_XLSX_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        for v in visits:
            promoter_name = (
                f"{v.promoter.first_name} {v.promoter.last_name}" if v.promoter else ""
            )
            ws.append(
                [
                    v.week_label,
                    v.date.isoformat(),
                    str(v.start_time)[:5],
                    str(v.end_time)[:5],
                    v.pos.cdb_code,
                    v.pos.name,
                    v.pos.city,
                    v.pos.priority,
                    promoter_name,
                    v.programme_type,
                    v.comments,
                ]
            )

        filename = f"{schedule.name}.xlsx"
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ScheduleImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, pk):
        schedule = get_object_or_404(Schedule, pk=pk)
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"error": "Could not read the file. Make sure it is a valid .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active

        pos_by_cdb = {p.cdb_code: p for p in PointOfSale.objects.all()}
        promoter_by_name = {
            f"{p.first_name} {p.last_name}": p for p in Promoter.objects.all()
        }

        schedule.visits.all().delete()
        created, errors = [], []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue  # skip blank rows

            try:
                (
                    week_label,
                    date_val,
                    start_val,
                    end_val,
                    cdb_code,
                    _pos_name,
                    _city,
                    _priority,
                    promoter_name,
                    programme_type,
                    comments,
                ) = (row + (None,) * 11)[:11]

                if not date_val or not cdb_code:
                    errors.append(f"Row {row_num}: missing date or CDB code — skipped.")
                    continue

                pos = pos_by_cdb.get(str(cdb_code).strip())
                if pos is None:
                    errors.append(
                        f"Row {row_num}: POS '{cdb_code}' not found — skipped."
                    )
                    continue

                date_obj = _parse_xlsx_date(date_val)
                if not (schedule.period_start <= date_obj <= schedule.period_end):
                    errors.append(
                        f"Row {row_num}: date {date_obj} outside schedule period"
                        " — skipped."
                    )
                    continue

                promoter = (
                    promoter_by_name.get(str(promoter_name).strip())
                    if promoter_name
                    else None
                )
                week_num = (date_obj - schedule.period_start).days // 7 + 1
                visit = ScheduledVisit.objects.create(
                    schedule=schedule,
                    pos=pos,
                    promoter=promoter,
                    date=date_obj,
                    start_time=_parse_xlsx_time(start_val) if start_val else "09:00",
                    end_time=_parse_xlsx_time(end_val) if end_val else "11:00",
                    programme_type=(
                        str(programme_type).strip()
                        if programme_type
                        else (promoter.programme_type if promoter else "Radical")
                    ),
                    week_label=(
                        str(week_label).strip() if week_label else f"W{week_num}"
                    ),
                    comments=str(comments).strip() if comments else "",
                )
                created.append(visit)

            except Exception as exc:
                errors.append(f"Row {row_num}: {exc}")

        serializer = ScheduledVisitSerializer(created, many=True)
        return Response(
            {"visits": serializer.data, "errors": errors},
            status=status.HTTP_200_OK,
        )
