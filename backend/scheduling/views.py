import datetime
import json

import openpyxl
from django.conf import settings
from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import generics, status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView

from .ai import generate_schedule, stream_generate_schedule
from .models import PointOfSale, Promoter, Schedule, ScheduledVisit
from .serializers import (
    PointOfSaleSerializer,
    PromoterSerializer,
    ScheduleCreateSerializer,
    ScheduledVisitSerializer,
    ScheduleSerializer,
)

# ── Shared helpers ─────────────────────────────────────────────────────────────

_XLSX_HEADERS = [
    "Week",
    "Date",
    "Start Time",
    "End Time",
    "CDB Code",
    "POS Name",
    "City",
    "Priority",
    "Promoter",
    "Programme",
    "AI Reasoning",
]
_XLSX_COL_WIDTHS = [8, 13, 11, 11, 13, 28, 16, 13, 28, 13, 55]


def _create_visits_from_ai(schedule, visit_data, pos_map, promoter_map):
    """
    Delete all existing visits for *schedule* and create new ones from
    the AI-generated *visit_data* list (dicts with pos_id, promoter_id, …).

    Returns (created_visits, errors).
    """
    schedule.visits.all().delete()
    created, errors = [], []

    for v in visit_data:
        pos = pos_map.get(v.get("pos_id"))
        if pos is None:
            errors.append(f"Unknown pos_id {v.get('pos_id')} — skipped.")
            continue
        promoter = promoter_map.get(v.get("promoter_id"))
        try:
            date_obj = datetime.date.fromisoformat(v["date"])
            week_num = (date_obj - schedule.period_start).days // 7 + 1
            visit = ScheduledVisit.objects.create(
                schedule=schedule,
                pos=pos,
                promoter=promoter,
                date=date_obj,
                start_time=v["start_time"],
                end_time=v["end_time"],
                programme_type=(promoter.programme_type if promoter else "Radical"),
                week_label=f"W{week_num}",
                comments=v.get("reason", ""),
            )
            created.append(visit)
        except Exception as exc:
            errors.append(str(exc))

    return created, errors


def _parse_xlsx_time(val) -> str:
    """Return HH:MM string from an openpyxl cell value."""
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M")
    return str(val).strip()[:5]


def _parse_xlsx_date(val) -> datetime.date:
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return datetime.date.fromisoformat(str(val).strip())


# ── Schedule list / create ─────────────────────────────────────────────────────


class ScheduleListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return ScheduleCreateSerializer
        return ScheduleSerializer

    def get_queryset(self):
        qs = Schedule.objects.select_related("created_by").order_by("-period_start")
        status_filter = self.request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        out = ScheduleSerializer(serializer.instance, context={"request": request})
        headers = self.get_success_headers(out.data)
        return Response(out.data, status=status.HTTP_201_CREATED, headers=headers)


# ── Schedule detail ────────────────────────────────────────────────────────────


class ScheduleDetailView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ScheduleSerializer
    queryset = Schedule.objects.select_related("created_by")


# ── Visit list ─────────────────────────────────────────────────────────────────


class ScheduleVisitListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ScheduledVisitSerializer

    def get_queryset(self):
        schedule = get_object_or_404(Schedule, pk=self.kwargs["pk"])
        return (
            ScheduledVisit.objects.filter(schedule=schedule)
            .select_related("pos", "promoter")
            .order_by("date", "start_time")
        )


# ── AI generation ──────────────────────────────────────────────────────────────


class ServerSentEventRenderer(BaseRenderer):
    media_type = "text/event-stream"
    format = "event-stream"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class ScheduleGenerateView(APIView):
    permission_classes = [IsAuthenticated]
    renderer_classes = [JSONRenderer, ServerSentEventRenderer]

    def post(self, request, pk):
        if not settings.OPENAI_API_KEY:
            return Response(
                {"error": "OPENAI_API_KEY is not configured on the server."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        schedule = get_object_or_404(
            Schedule.objects.prefetch_related("included_pos", "included_promoters"),
            pk=pk,
        )
        optimization_goal = request.data.get(
            "optimization_goal", "sales * 10 + interviews"
        )
        user_prompt = request.data.get("user_prompt", "")

        if "text/event-stream" in request.META.get("HTTP_ACCEPT", ""):
            return self._stream(schedule, optimization_goal, user_prompt)

        # Blocking path — used by the Django test suite (mocked at the view level)
        try:
            result = generate_schedule(schedule, optimization_goal, user_prompt)
        except Exception as e:
            return Response(
                {"error": f"AI generation failed: {e}"},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        pos_map = {p.id: p for p in schedule.included_pos.all()}
        promoter_map = {p.id: p for p in schedule.included_promoters.all()}
        created, errors = _create_visits_from_ai(
            schedule, result["visits"], pos_map, promoter_map
        )

        schedule.score = result.get("score")
        schedule.save(update_fields=["score"])

        serializer = ScheduledVisitSerializer(created, many=True)
        return Response(
            {
                "summary": result["summary"],
                "score": result.get("score"),
                "visits": serializer.data,
                "usage": result["usage"],
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )

    def _stream(self, schedule, optimization_goal, user_prompt):
        pos_map = {p.id: p for p in schedule.included_pos.all()}
        promoter_map = {p.id: p for p in schedule.included_promoters.all()}

        def _event_generator():
            for event in stream_generate_schedule(
                schedule, optimization_goal, user_prompt
            ):
                if event["type"] in ("thinking", "error"):
                    yield f"data: {json.dumps(event)}\n\n"
                    continue

                # "done" — persist visits + score, then emit the SSE payload
                created, errors = _create_visits_from_ai(
                    schedule, event.get("visits", []), pos_map, promoter_map
                )
                schedule.score = event.get("score")
                schedule.save(update_fields=["score"])

                serializer = ScheduledVisitSerializer(created, many=True)
                payload = {
                    "type": "done",
                    "summary": event.get("summary", ""),
                    "score": event.get("score"),
                    "visits": list(serializer.data),
                    "usage": event.get("usage", {}),
                    "errors": errors,
                }
                yield f"data: {json.dumps(payload, default=str)}\n\n"

        response = StreamingHttpResponse(
            _event_generator(), content_type="text/event-stream"
        )
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        return response


# ── Export ─────────────────────────────────────────────────────────────────────


class ScheduleExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        schedule = get_object_or_404(Schedule, pk=pk)
        visits = (
            ScheduledVisit.objects.filter(schedule=schedule)
            .select_related("pos", "promoter")
            .order_by("date", "start_time")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = schedule.name[:31]  # Excel sheet name limit
        ws.freeze_panes = "A2"

        ws.append(_XLSX_HEADERS)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        for col, width in enumerate(_XLSX_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        for v in visits:
            promoter_name = (
                f"{v.promoter.first_name} {v.promoter.last_name}" if v.promoter else ""
            )
            ws.append(
                [
                    v.week_label,
                    v.date.isoformat(),
                    str(v.start_time)[:5],
                    str(v.end_time)[:5],
                    v.pos.cdb_code,
                    v.pos.name,
                    v.pos.city,
                    v.pos.priority,
                    promoter_name,
                    v.programme_type,
                    v.comments,
                ]
            )

        filename = f"{schedule.name}.xlsx"
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ── Import ─────────────────────────────────────────────────────────────────────


class ScheduleImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, pk):
        schedule = get_object_or_404(Schedule, pk=pk)
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"error": "Could not read the file. Make sure it is a valid .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active

        pos_by_cdb = {p.cdb_code: p for p in PointOfSale.objects.all()}
        promoter_by_name = {
            f"{p.first_name} {p.last_name}": p for p in Promoter.objects.all()
        }

        schedule.visits.all().delete()
        created, errors = [], []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue  # skip blank rows

            try:
                (
                    week_label,
                    date_val,
                    start_val,
                    end_val,
                    cdb_code,
                    _pos_name,
                    _city,
                    _priority,
                    promoter_name,
                    programme_type,
                    comments,
                ) = (row + (None,) * 11)[:11]

                if not date_val or not cdb_code:
                    errors.append(f"Row {row_num}: missing date or CDB code — skipped.")
                    continue

                pos = pos_by_cdb.get(str(cdb_code).strip())
                if pos is None:
                    errors.append(
                        f"Row {row_num}: POS '{cdb_code}' not found — skipped."
                    )
                    continue

                date_obj = _parse_xlsx_date(date_val)
                if not (schedule.period_start <= date_obj <= schedule.period_end):
                    errors.append(
                        f"Row {row_num}: date {date_obj} outside schedule period"
                        " — skipped."
                    )
                    continue

                promoter = (
                    promoter_by_name.get(str(promoter_name).strip())
                    if promoter_name
                    else None
                )
                week_num = (date_obj - schedule.period_start).days // 7 + 1
                visit = ScheduledVisit.objects.create(
                    schedule=schedule,
                    pos=pos,
                    promoter=promoter,
                    date=date_obj,
                    start_time=_parse_xlsx_time(start_val) if start_val else "09:00",
                    end_time=_parse_xlsx_time(end_val) if end_val else "11:00",
                    programme_type=(
                        str(programme_type).strip()
                        if programme_type
                        else (promoter.programme_type if promoter else "Radical")
                    ),
                    week_label=(
                        str(week_label).strip() if week_label else f"W{week_num}"
                    ),
                    comments=str(comments).strip() if comments else "",
                )
                created.append(visit)

            except Exception as exc:
                errors.append(f"Row {row_num}: {exc}")

        serializer = ScheduledVisitSerializer(created, many=True)
        return Response(
            {"visits": serializer.data, "errors": errors},
            status=status.HTTP_200_OK,
        )


# ── Publish ────────────────────────────────────────────────────────────────────


class SchedulePublishView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        schedule = get_object_or_404(
            Schedule.objects.select_related("created_by"), pk=pk
        )
        if schedule.status != Schedule.Status.DRAFT:
            return Response(
                {"error": "Only Draft schedules can be published."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        schedule.status = Schedule.Status.PUBLISHED
        schedule.save(update_fields=["status"])
        return Response(ScheduleSerializer(schedule).data)


# ── POS / Promoter lists ───────────────────────────────────────────────────────


class PointOfSaleListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PointOfSaleSerializer
    queryset = PointOfSale.objects.filter(is_active=True).order_by("name")


class PromoterListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PromoterSerializer
    queryset = Promoter.objects.filter(is_active=True).order_by(
        "last_name", "first_name"
    )
